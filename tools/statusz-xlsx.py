#!/usr/bin/env python3
"""
statusz-xlsx.py — a státuszjelentés átalakítása színezett Excellé

MIÉRT KÜLÖN LÉPÉS
A statusz.mjs Node-ban fut, és a JSON/CSV kiírásához nincs
szüksége külső csomagra. A színezett Excelhez viszont openpyxl
kell, ami Python. Külön lépésként ez tiszta felosztás: a
gyűjtés és a megjelenítés nem keveredik.

MIT AD
Egy munkalap, sorokban a pályákkal, színkóddal:
    zöld   – van friss versenynyom
    sárga  – tisztázatlan, kézi ellenőrzés kell
    piros  – bezárásra utaló bizonyíték

A "Döntés" oszlop üres, legördülő listával: oda írod be, mit
fogadsz el. Ez a fájl a MUNKAESZKÖZ, nem a végeredmény – a
tényleges adatbázis-módosítás továbbra is kézi.

Használat:
    python3 tools/statusz-xlsx.py
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ITT = os.path.dirname(os.path.abspath(__file__))
BE = os.path.join(ITT, 'statusz-jelentes.json')
KI = os.path.join(ITT, 'statusz-jelentes.xlsx')

# A dokumentumban végig ugyanaz a betűtípus, hogy nyomtatva is
# egységes legyen.
BETU = 'Arial'

SZINEK = {
    'zöld':  'C6EFCE',   # halvány zöld
    'sárga': 'FFEB9C',   # halvány sárga
    'piros': 'FFC7CE',   # halvány piros
}
SZOVEGSZIN = {
    'zöld':  '006100',
    'sárga': '7F6000',
    'piros': '9C0006',
}

OSZLOPOK = [
    ('Pálya',         34),
    ('Település',     22),
    ('Jelenlegi',     11),
    ('Javasolt',      11),
    ('Eltérés',        9),
    ('Utolsó nyom',   13),
    ('Családok',       9),
    ('Mely források',  30),
    ('Saját honlap',  18),
    ('Indok',         52),
    ('DÖNTÉS',        14),
    ('Megjegyzés',    30),
]


def main():
    if not os.path.exists(BE):
        print(f'Nem található: {BE}', file=sys.stderr)
        print('Előbb futtasd: node tools/statusz.mjs --orszag AUS', file=sys.stderr)
        sys.exit(1)

    with open(BE, encoding='utf-8') as f:
        adat = json.load(f)

    eredmenyek = adat.get('eredmenyek', [])
    if not eredmenyek:
        print('A jelentés üres.', file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Státusz {adat.get('iso', '')}"

    # ---- Fejléc-sáv: mit tartalmaz a fájl -------------------
    ws['A1'] = f"Státuszellenőrzés — {adat.get('iso','')}"
    ws['A1'].font = Font(name=BETU, size=13, bold=True)

    ws['A2'] = (f"Vizsgált időszak: {adat.get('honapok','?')} hónap · "
                f"Pálya: {len(eredmenyek)} · "
                f"Elérhető forrás: {len(adat.get('hasznaltForrasok', []))}")
    ws['A2'].font = Font(name=BETU, size=9, italic=True, color='666666')

    ws['A3'] = ('Ez JAVASLAT, nem döntés. A hiányzás nem bizonyítja a bezárást: '
                'egy pálya kimaradhat idényen kívül vagy felújítás alatt. '
                'A DÖNTÉS oszlopba írd be, mit fogadsz el.')
    ws['A3'].font = Font(name=BETU, size=9, color='9C0006')

    FEJ = 5   # a táblázat fejlécsora

    # ---- Fejléc --------------------------------------------
    vekony = Side(style='thin', color='D0D0D0')
    keret = Border(left=vekony, right=vekony, top=vekony, bottom=vekony)

    for i, (cim, szel) in enumerate(OSZLOPOK, start=1):
        c = ws.cell(row=FEJ, column=i, value=cim)
        c.font = Font(name=BETU, size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='3E2318')   # a márka barnája
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = keret
        ws.column_dimensions[get_column_letter(i)].width = szel

    # ---- Sorok ---------------------------------------------
    # Rendezés: előbb ami döntést igényel (piros, sárga), és azon
    # belül ami eltér a jelenlegi státusztól. Így a lista tetején
    # van a munka, nem kell görgetni érte.
    rang = {'piros': 0, 'sárga': 1, 'zöld': 2}
    eredmenyek.sort(key=lambda e: (
        rang.get(e.get('szin'), 3),
        0 if e.get('javaslat') != e.get('jelenlegi') else 1,
        e.get('nev', '')
    ))

    for sor, e in enumerate(eredmenyek, start=FEJ + 1):
        szin = e.get('szin', 'sárga')
        elter = e.get('javaslat') != e.get('jelenlegi')

        ertekek = [
            e.get('nev'),
            e.get('varos'),
            e.get('jelenlegi'),
            e.get('javaslat'),
            'IGEN' if elter else '',
            e.get('legutobbi') or '—',
            len(e.get('csaladok', [])),
            ', '.join(e.get('csaladok', [])),
            e.get('sajatHonlap') or '—',
            e.get('indok'),
            '',    # DÖNTÉS – kézzel
            '',    # Megjegyzés – kézzel
        ]

        for i, ertek in enumerate(ertekek, start=1):
            c = ws.cell(row=sor, column=i, value=ertek)
            c.font = Font(name=BETU, size=10,
                          bold=(i == 1),
                          color=SZOVEGSZIN.get(szin, '000000'))
            c.fill = PatternFill('solid', fgColor=SZINEK.get(szin, 'FFFFFF'))
            c.border = keret
            c.alignment = Alignment(
                vertical='top',
                wrap_text=(i in (1, 2, 8, 10)),
                horizontal='center' if i in (3, 4, 5, 6, 7) else 'left'
            )

    # ---- A DÖNTÉS oszlop: legördülő lista -------------------
    # Így nem lehet elgépelni a státuszt, és később gépi
    # feldolgozásra is alkalmas marad.
    dv = DataValidation(
        type='list',
        formula1='"active,inactive,unknown,closed,MARAD,ELLENORIZNI"',
        allow_blank=True,
        showDropDown=False
    )
    dv.error = 'Válassz a listából.'
    dv.errorTitle = 'Érvénytelen érték'
    ws.add_data_validation(dv)
    dv.add(f'K{FEJ+1}:K{FEJ+len(eredmenyek)}')

    # ---- Kezelhetőség --------------------------------------
    ws.freeze_panes = f'A{FEJ+1}'          # a fejléc mindig látszik
    ws.auto_filter.ref = f'A{FEJ}:L{FEJ+len(eredmenyek)}'

    # ---- Második lap: a felhasznált források ----------------
    ws2 = wb.create_sheet('Források')
    ws2['A1'] = 'Felhasznált források'
    ws2['A1'].font = Font(name=BETU, size=12, bold=True)
    ws2['A2'] = ('Csak azok szerepelnek, amelyek ténylegesen elérhetők voltak. '
                 'Amit a robots.txt tiltott vagy hibára futott, kimaradt.')
    ws2['A2'].font = Font(name=BETU, size=9, italic=True, color='666666')
    ws2.column_dimensions['A'].width = 60

    for i, nev in enumerate(adat.get('hasznaltForrasok', []), start=4):
        ws2.cell(row=i, column=1, value=nev).font = Font(name=BETU, size=10)

    # A napló is ide kerül – így nem terheli a futás kimenetét,
    # de utólag visszakereshető, mi miért maradt ki.
    naplo = adat.get('naplo', [])
    if naplo:
        s = 4 + len(adat.get('hasznaltForrasok', [])) + 2
        ws2.cell(row=s, column=1, value='Részletes napló').font = \
            Font(name=BETU, size=11, bold=True)
        for i, x in enumerate(naplo, start=s + 1):
            ws2.cell(row=i, column=1, value=x.strip()).font = Font(name=BETU, size=9)

    wb.save(KI)

    piros = sum(1 for e in eredmenyek if e.get('szin') == 'piros')
    sarga = sum(1 for e in eredmenyek if e.get('szin') == 'sárga')
    zold = sum(1 for e in eredmenyek if e.get('szin') == 'zöld')
    print(f'Excel kész: {os.path.basename(KI)}')
    print(f'  piros {piros} · sárga {sarga} · zöld {zold}')


if __name__ == '__main__':
    main()
